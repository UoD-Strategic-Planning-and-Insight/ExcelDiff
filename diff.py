"""
Contains the TableDiff class, for processing the differences between Excel tables, and associated supporting classes.
"""

from dataclasses import dataclass
from typing import Any

import openpyxl
from openpyxl.cell import Cell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table

import Utils

from xltables import XLTable


@dataclass
class TableReference:
    """A reference by filepath to an Excel file, and to a specific table within that Excel file."""

    filepath: str
    """The filepath of the Excel file containing the table."""

    sheet_name: str
    """The name of the sheet in the Excel file that contains the table."""

    table_name: str
    """The name of the table."""


@dataclass
class CellDifference:
    """
    A record of a difference between two cells of the same column in different rows, which may not necessarily be in the
    same table.
    """

    column_name: str
    """The name of the column the cells are in."""

    value1: Any
    """The value of the cell in the first row."""

    value2: Any
    """The value of the cell in the second row."""


@dataclass
class RowDifference:
    """A record of a set of differences between two rows with the same keys in different tables."""

    keys: dict[str, Any]
    """
    A dictionary of the shared keys between the two rows; the dictionary key is the column name and the dictionary value
    is the value of the corresponding key cell in those rows. This is a dictionary rather than just a pair of values,
    because tables may have compound keys - each key of this dictionary would be the name of one of the columns that
    make up the compound key.
    """

    cell_differences: list[CellDifference]
    """A list of cell differences between the two rows."""


@dataclass
class TableColumnContent:
    """A record of the values of cells in a column, along with the column name."""

    column_name: str
    """The name of the column."""

    values: list[Any]
    """A list of the values of the cells in the represented column."""


class TableDiff:
    """
    A queued difference between two tables.

    The difference isn't processed immediately. Rather, this contains the information required to establish the
    differences between two tables, and then save them to a particular file. Once you have a constructed instance of
    this class, you can establish the differences between two tables and save them to a file with `.process_and_save()`.

    The file produced details the differences between rows shared between the two tables, rows that only exist in one
    table or the other, and columns that only exist in one table or the other.
    """

    # TODO: Make this store the rows unique to one table or the other as dictionaries of values, rather than of cells,
    #       so that there's no possible problem programmatically reading those values once the tables have been
    #       discarded.

    # TODO: Add a method to process a difference without saving it to a file, so this can be used without saving
    #       necessarily anything. When done, update inline documentation that directs users to a method to process this
    #       diff, to mention that method as well.


    first_table_ref:  TableReference
    """A reference to one of the tables being compared."""

    second_table_ref: TableReference
    """A reference to the other table being compared."""

    result_filepath:  str
    """The filepath at which to save the differences."""

    key_column_names: list[str]
    """
    The names of the columns present in both tables that collectively form a unique identifier. This allows the contents
    of tables to be compared without having to worry about order.
    """


    first_table: XLTable | None
    """One of the tables being compared."""

    second_table: XLTable | None
    """The other table being compared."""


    row_numbers_for_key_sets_in_first:  dict[str, int]
    """The numbers of every row in the first table, mapped against a the key values of that row, encoded as a string."""

    row_numbers_for_key_sets_in_second: dict[str, int]
    """
    The numbers of every row in the second table, mapped against a the key values of that row, encoded as a string.
    """


    row_differences:        list[RowDifference]
    """A list of the different rows between the two tables. Only available once processed."""

    rows_only_in_first:     list[dict[str, Cell]]
    """A list of the rows that only exist in the first table. Only available once processed."""

    rows_only_in_second:    list[dict[str, Cell]]
    """A list of the rows that only exist in the second table. Only available once processed."""

    columns_only_in_first:  list[TableColumnContent]
    """A list of the columns that only exist in the first table. Only available once processed."""

    columns_only_in_second: list[TableColumnContent]
    """A list of the columns that only exist in the second table. Only available once processed."""

    def __init__(self,
                 first:            TableReference,
                 second:           TableReference,
                 result_filepath:  str,
                 key_column_names: list[str]):
        """
        Creates a new TableDiff object.

        This does not immediately process the difference. To process the difference, call `.process_and_save()`
        :param first: A reference to the first table being compared.
        :param second: A reference to the second table being compared.
        :param result_filepath: The filepath the resulting table should be saved to.
        :param key_column_names: The names of the columns common to both tables that collectively form a
                                 uniquely-identifying key. This will not behave properly if the given key is not
                                 completely unique to each row.
        """

        self.first_table_ref  = first
        self.second_table_ref = second
        self.result_filepath  = result_filepath
        self.key_column_names = key_column_names

        self.row_numbers_for_key_sets_in_first  = {}
        self.row_numbers_for_key_sets_in_second = {}

        self.row_differences        = []
        self.rows_only_in_first     = []
        self.rows_only_in_second    = []
        self.columns_only_in_first  = []
        self.columns_only_in_second = []

    def process_and_save(self) -> None:
        """
        Processes the differences between the two tables in this diff, and saves those differences to an Excel file at
        the filepath stored.

        After calling this, information about the differences between the two tables will be available in this object.
        """

        self.load_tables()
        self.build_table_indices()
        self.read_row_differences()
        self.read_rows_only_in_second()
        self.read_columns_only_in_first()
        self.read_columns_only_in_second()
        self.save_to_file()
        self.discard_loaded_tables()

    def load_tables(self) -> None:
        """
        Loads the tables referenced by this diff.
        """

        ref1              = self.first_table_ref
        ref2              = self.second_table_ref
        self.first_table  = XLTable.load_from_file(ref1.filepath, ref1.sheet_name, ref1.table_name)
        self.second_table = XLTable.load_from_file(ref2.filepath, ref2.sheet_name, ref2.table_name)

    def discard_loaded_tables(self) -> None:
        """
        Closes and discards the tables referenced by this diff.
        """

        self.first_table.source_workbook.close()
        self.second_table.source_workbook.close()
        self.first_table  = None
        self.second_table = None

    def build_table_indices(self) -> None:
        """
        Builds indexes of the loaded tables, of the keys for each row against their row numbers. This allows for faster
        random access to rows.
        """

        self._build_row_index(self.first_table,  self.row_numbers_for_key_sets_in_first)
        self._build_row_index(self.second_table, self.row_numbers_for_key_sets_in_second)

    def read_row_differences(self) -> None:
        """
        Reads the differences between rows common to both tables into this object.

        For speed's sake, this also reads the rows unique to the first table into this object.
        """

        self.row_differences    = []
        self.rows_only_in_first = []

        for row in self.first_table.row_iterator:
            keys: dict[str, Any] = {}

            for key_col_name in self.key_column_names:
                keys[key_col_name] = row[key_col_name].value

            matching_row_in_second: dict[str, Cell] | None \
                = self._get_row_with_keys(self.second_table, keys, self.row_numbers_for_key_sets_in_second)

            if(matching_row_in_second is None):
                self.rows_only_in_first.append(row)
                continue

            cell_diffs: list[CellDifference] = self._get_differences_between_rows(row, matching_row_in_second)

            if(len(cell_diffs) != 0):
                self.row_differences.append(RowDifference(keys, cell_diffs))

    def read_rows_only_in_second(self) -> None:
        """
        Reads the rows unique to the second table into this object.
        """

        self.rows_only_in_second = []

        for row in self.second_table.row_iterator:
            keys: dict[str, Any] = {}

            for key_col_name in self.key_column_names:
                keys[key_col_name] = row[key_col_name].value

            matching_row_exists_in_first: bool \
                = self.row_numbers_for_key_sets_in_first.get(Utils.dict_to_str(keys)) is not None

            if(not matching_row_exists_in_first):
                self.rows_only_in_second.append(row)

    def read_columns_only_in_first(self) -> None:
        """
        Reads the rows unique to the first table into this object.
        """

        self.columns_only_in_first = self._get_columns_not_in_other(self.first_table, self.second_table)

    def read_columns_only_in_second(self) -> None:
        """
        Reads the rows unique to the second table into this object.
        """

        self.columns_only_in_second = self._get_columns_not_in_other(self.second_table, self.first_table)

    def save_to_file(self) -> None:
        """
        Creates an Excel file at the stored filepath and populates it, as needed, with sheets for the differences
        between common rows, the rows unique to one table or another, and the columns unique to one table or another.
        """

        wb = openpyxl.Workbook()

        self._add_diffs_sheet_to_workbook(wb)
        self._add_rows_only_in_one_sheet_to_workbook(wb, self.rows_only_in_first,
                                                     2, "Rows unique to first", "RowsUniqueToFirst")

        self._add_rows_only_in_one_sheet_to_workbook(wb, self.rows_only_in_second,
                                                     3, "Rows unique to second", "RowsUniqueToSecond")

        key_cols_in_first:  list[TableColumnContent] = self._get_key_columns(self.first_table)
        key_cols_in_second: list[TableColumnContent] = self._get_key_columns(self.second_table)

        self._add_columns_only_in_one_sheet_to_workbook(wb, key_cols_in_first, self.columns_only_in_first,
                                                        4, "Columns unique to first", "ColumnsUniqueToFirst")

        self._add_columns_only_in_one_sheet_to_workbook(wb, key_cols_in_second, self.columns_only_in_second,
                                                        5, "Columns unique to second", "ColumnsUniqueToSecond")

        wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
        wb.save(self.result_filepath)

    def _add_diffs_sheet_to_workbook(self, wb: Workbook) -> None:
        """
        Write the differences between common rows that have been processed into the given workbook as a sheet.
        :param wb: The workbook to write the sheet into.
        """

        if(len(self.row_differences) == 0):
            return

        wb.create_sheet("Differences", 1)
        sheet = wb.get_sheet_by_name("Differences")

        for i in range(len(self.key_column_names)):
            sheet.cell(1, i + 1).value = self.key_column_names[i]

        table = Table(displayName = "DiffTable",
                      ref         = f"A1:{Utils.convert_int_to_alphabetic_number(len(self.key_column_names))}1")

        sheet.add_table(table)
        tbl_diff = XLTable(self.result_filepath, wb, sheet, table)

        for diff in self.row_differences:
            tbl_diff.add_row()
            row = tbl_diff.bottom_row

            for k, v in diff.keys.items():
                row[k].value = v

            for cell_diff in diff.cell_differences:
                col_name_1 = cell_diff.column_name + " * 1"
                col_name_2 = cell_diff.column_name + " * 2"
                columns_added: bool = False

                if(not tbl_diff.has_column(col_name_1)):
                    tbl_diff.add_column(col_name_1)
                    tbl_diff.add_column(col_name_2)
                    columns_added = True

                if(columns_added):
                    row = tbl_diff.bottom_row

                row[col_name_1].value = cell_diff.value1
                row[col_name_2].value = cell_diff.value2

    def _add_rows_only_in_one_sheet_to_workbook(self,
                                                wb:          Workbook,
                                                rows:        list[dict[str, Cell]],
                                                sheet_index: int,
                                                sheet_name:  str,
                                                table_name:  str) \
            -> None:
        """
        Write the rows unique to one of the tables to the given workbook as a sheet.
        :param wb: The workbook to write the sheet into.
        :param rows: The rows unique to the table.
        :param sheet_index: The index of the sheet in the workbook.
        :param sheet_name: The name of the sheet.
        :param table_name: The name of the table to be written.
        """

        if(len(rows) == 0):
            return

        Workbook.create_sheet(wb, sheet_name, sheet_index)
        sheet: Worksheet = wb.get_sheet_by_name(sheet_name)
        key_column_count = len(self.key_column_names)

        for i in range(key_column_count):
            sheet.cell(1, i + 1).value = self.key_column_names[i]

        non_key_column_names = [x for x in rows[0].keys() if x not in self.key_column_names]
        table_width = len(rows[0].keys())

        for i in range(len(non_key_column_names)):
            sheet.cell(1, key_column_count + i + 1).value = non_key_column_names[i]

        table = Table(displayName = table_name,
                      ref         = f"A1:{Utils.convert_int_to_alphabetic_number(table_width)}1")

        sheet.add_table(table)
        tbl = XLTable(self.result_filepath, wb, sheet, table)

        for source_row in rows:
            tbl.add_row()
            dest_row = tbl.bottom_row

            for k, v in source_row.items():
                dest_row[k].value = v.value

    def _add_columns_only_in_one_sheet_to_workbook(self,
                                                   wb:          Workbook,
                                                   key_columns: list[TableColumnContent],
                                                   columns:     list[TableColumnContent],
                                                   sheet_index: int,
                                                   sheet_name:  str,
                                                   table_name:  str) \
            -> None:
        """
        Write the columns unique to one of the tables to the given workbook as a sheet.
        :param wb: The workbook to write the sheet into.
        :param key_columns: The columns used to uniquely identify rows in the two tables.
        :param columns: The columns unique to one of the tables.
        :param sheet_index: The index of the sheet in the workbook.
        :param sheet_name: The name of the sheet.
        :param table_name: The name of the table to be written.
        """

        if(len(columns) == 0):
            return

        wb.create_sheet(sheet_name, sheet_index)
        sheet = wb.get_sheet_by_name(sheet_name)
        row_count = len(columns[0].values)

        for i in range(len(key_columns)):
            col: TableColumnContent = key_columns[i]
            sheet.cell(1, i + 1).value = col.column_name

            for j in range(len(col.values)):
                sheet.cell(j + 2, i + 1).value = col.values[j]

        table = Table(displayName=table_name,
                      ref=f"A1:{Utils.convert_int_to_alphabetic_number(len(key_columns))}{row_count + 1}")

        sheet.add_table(table)
        tbl = XLTable(self.result_filepath, wb, sheet, table)

        for col in columns:
            tbl.add_column(col.column_name, col.values)

    def _build_row_index(self, table: XLTable, index: dict[str, int]) -> None:
        """
        Populates a given dictionary with string-encoded versions of the keys of every row, and the number that row
        appears in.
        :param table: The table this is an index for.
        :param index: The dictionary serving as an index. It should be empty.
        """

        row_no: int = -1

        for row in table.row_iterator:
            keys: dict[str, str] = {}
            row_no += 1

            for key_col_name in self.key_column_names:
                keys[key_col_name] = row[key_col_name].value

            key_str: str = Utils.dict_to_str(keys)
            index[key_str] = row_no

    def _get_key_columns(self, table: XLTable) -> list[TableColumnContent]:
        """
        Gets a list of the key columns in full (their names and contents) from the given table.
        :param table: The table to get the key columns from.
        :return: A list of the columns (as TableColumnContent objects) that form the uniquely identifying key in the
                 given table.
        """

        result: list[TableColumnContent] = []

        for col_name in self.key_column_names:
            col_vals: list[Any] = []

            for row in table.row_iterator:
                col_vals.append(row[col_name].value)

            result.append(TableColumnContent(col_name, col_vals))

        return result


    def _get_row_with_keys(self, table: XLTable, keys: dict[str, Any], row_number_lookup_dict: dict[str, int])\
            -> dict[str, Cell] | None:
        """
        Gets the row in the given table with the given keys.
        :param table: The table to look a row up in.
        :param keys: Dictionary where the keys are the names of columns that make up part of the uniquely identifying
                     key in the table, and the values are the values for those columns in the sought-after row.
        :param row_number_lookup_dict: A dictionary containing the numbers of every row in the given table, mapped to
                                       the keys for those rows encoded as a string.
        :return: If a row was found in the table (using the given index), that row as a dictionary of cells mapped to
                 their column names. Otherwise, null.
        """

        key_string: str = Utils.dict_to_str(keys)
        row_no: int | None = row_number_lookup_dict.get(key_string)
        return (table.get_row(row_no)) if (row_no is not None) else (None)

    def _get_differences_between_rows(self, first: dict[str, Cell], second: dict[str, Cell]) \
            -> list[CellDifference]:
        """
        Gets the differences between two rows.
        :param first: One of the rows to compare, as a dictionary where the keys are the column names and the values are
                      the corresponding cells.
        :param second: The other row to compare, as a dictionary where the keys are the column names and the values are
                      the corresponding cells.
        :return: A list of cell differences, differences between cells in the given rows from the same columns.
        """

        result: list[CellDifference] = []

        for k, v1 in first.items():
            v2: Cell | None = second.get(k)

            if(v2 is None):
                continue

            v1val = str(v1.value).strip() if v1.value is not None else ""
            v2val = str(v2.value).strip() if v2.value is not None else ""

            if(v1val != v2val):
                result.append(CellDifference(k, v1val, v2val))

        return result

    def _get_columns_not_in_other(self, table: XLTable, other_table: XLTable) \
            -> list[TableColumnContent]:
        """
        Gets the columns in unique to one of the tables.
        :param table: The table that may contain columns not in the other.
        :param other_table: The other table to compare.
        :return: A list of the columns in the first table that are not present in the second.
        """

        col_names_1 = table.column_names
        col_names_2 = other_table.column_names

        cols_not_in_other: list[TableColumnContent] = []

        for col_name in col_names_1:
            if(col_name not in col_names_2):
                col_vals: list[Any] = []

                for row in table.row_iterator:
                    col_vals.append(row[col_name].value)

                cols_not_in_other.append(TableColumnContent(col_name, col_vals))

        return cols_not_in_other
