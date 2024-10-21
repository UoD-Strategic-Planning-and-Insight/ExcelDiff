import os.path
from typing import Any, List

import tkinter as tk

from tkinter import Tk, Label, Button, Grid, Listbox, Frame, LabelFrame, filedialog, Entry
from tkinter.font import Font
from tkinter.ttk import OptionMenu, Combobox

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet

from diff import TableDiff, TableReference


# TODO: Note: When showing a queue of diffs to process, if there is a first, second, or destination file chosen, present
#             a confirmation dialogue to make sure the user hasn't clicked the "start" button without finishing adding
#             the current diff they're in the process of adding.

# TODO: Add a progress bar when processing the diffs.

class MainWindow:
    # region UI fields

    _ui_window: Tk

    _ui_first_file_button: Button

    _ui_second_file_button: Button

    _ui_first_file_label: Label

    _ui_second_file_label: Label

    _ui_destination_file_label: Label

    _ui_first_file_table_menu: Combobox

    _ui_second_file_table_menu: Combobox

    _ui_key_list_menu: Combobox

    _ui_key_list: Listbox

    _ui_key_delete_button: Button

    _ui_button_row_without_queue: Frame | None = None

    _ui_enqueue_button: Button | None = None

    _ui_create_diff_button: Button | None = None

    _ui_diff_queue: Frame | None = None

    _ui_add_to_queue_button: Button | None = None

    _ui_diff_queue_listbox: Listbox | None = None

    _ui_remove_from_queue_button: Button | None = None

    _ui_create_diffs_from_queue_button: Button | None = None

    # endregion

    first_file_path: str | None = None

    second_file_path: str | None = None

    destination_file_path: str | None = None

    tables_in_first_file: list[TableReference] | None = None

    tables_in_second_file: list[TableReference] | None = None

    table_selected_from_first_file: TableReference | None = None

    table_selected_from_second_file: TableReference | None = None

    column_names_in_tables_from_first_file: dict[str, list[str]] | None = None

    column_names_in_tables_from_second_file: dict[str, list[str]] | None = None

    key_column_names: list[str] = []

    diff_queue: list[TableDiff] = []

    def display(self):
        window: Tk = Tk()
        self._ui_window = window
        window.geometry("400x420")
        window.wm_title("ExcelDiff")
        window.wm_minsize(400, 420)

        main_label_row = Frame(window)
        main_label_row.pack_configure(side="top", fill="x", pady=10)
        main_label = Label(main_label_row, text="Choose excel files to diff", font=Font(size=24), justify="center")
        main_label_row.bind("<Configure>", lambda e: main_label.config(wraplength=main_label_row.winfo_width()))
        main_label.pack_configure(fill="x", expand=True)

        file_chooser_button_row_container: Frame = Frame(window)
        file_chooser_button_row_container.pack_configure(side="top", fill="x")
        file_chooser_button_row: Frame = Frame(file_chooser_button_row_container)
        file_chooser_button_row.pack_configure(fill="x", expand=True)
        file_chooser_button_row.grid_columnconfigure(index=0, weight=1)
        file_chooser_button_row.grid_columnconfigure(index=1, weight=1)
        file_chooser_button_row_spacing = 5

        first_file_button: Button = Button(file_chooser_button_row, text="First", command=self.on_click_choose_first_file)
        first_file_button.grid_configure(column=0, row=0, sticky="EW", padx=file_chooser_button_row_spacing)
        self._ui_first_file_button = first_file_button
        first_file_path_label: Label = Label(file_chooser_button_row, text="(No file selected)")
        first_file_path_label.grid_configure(column=0, row=1, padx=file_chooser_button_row_spacing)
        self._ui_first_file_label = first_file_path_label
        first_file_table_menu: Combobox = Combobox(file_chooser_button_row, state="readonly")
        first_file_table_menu.grid_configure(column=0, row=2, sticky="EW", padx=file_chooser_button_row_spacing)
        first_file_table_menu.set("Choose a table...")
        self._ui_first_file_table_menu = first_file_table_menu
        first_file_table_menu.bind("<<ComboboxSelected>>", lambda e: self.on_choose_first_table())

        second_file_button: Button = Button(file_chooser_button_row, text="Second", command=self.on_click_choose_second_file)
        second_file_button.grid_configure(column=1, row=0, sticky="EW", padx=file_chooser_button_row_spacing)
        self._ui_second_file_button = second_file_button
        second_file_path_label: Label = Label(file_chooser_button_row, text="(No file selected)")
        second_file_path_label.grid_configure(column=1, row=1, padx=file_chooser_button_row_spacing)
        self._ui_second_file_label = second_file_path_label
        second_file_table_menu: Combobox = Combobox(file_chooser_button_row, state="readonly")
        second_file_table_menu.grid_configure(column=1, row=2, sticky="EW", padx=file_chooser_button_row_spacing)
        second_file_table_menu.set("Choose a table...")
        self._ui_second_file_table_menu = second_file_table_menu
        second_file_table_menu.bind("<<ComboboxSelected>>", lambda e: self.on_choose_second_table())

        diff_config_row_container: Frame = Frame(window)
        diff_config_row_container.pack_configure(side="top", fill="both", expand=True, pady=(20, 0))
        diff_config_row: Frame = Frame(diff_config_row_container)
        diff_config_row.pack_configure(fill="both", expand=True)
        diff_config_row.grid_columnconfigure(index=0, weight=1)
        diff_config_row.grid_columnconfigure(index=1, weight=1)
        diff_config_row.grid_rowconfigure(index=0, weight=1)
        diff_config_row_spacing = 5

        key_list_container: Frame = Frame(diff_config_row)
        key_list_container.grid_configure(column=0, row=0, sticky="NSEW", padx=diff_config_row_spacing, pady=diff_config_row_spacing)
        key_list_container.grid_columnconfigure(index=0, weight=1)
        key_list_container.rowconfigure(index=2, weight=1)
        key_list_label: Label = Label(key_list_container, text="Enter a key or compound key column names")
        key_list_label.grid_configure(column=0, row=0, sticky="EW")
        key_list_menu: Combobox = Combobox(key_list_container, state="disabled")
        key_list_menu.grid_configure(column=0, row=1, sticky="EW")
        key_list_menu.set("Choose a column...")
        key_list_menu.bind("<<ComboboxSelected>>", lambda e: self.on_choose_key_column())
        self._ui_key_list_menu = key_list_menu
        key_list_list: Listbox = Listbox(key_list_container, height=0)
        key_list_list.grid_configure(column=0, row=2, sticky="NSEW", pady=diff_config_row_spacing)
        key_list_list.bind("<<ListboxSelect>>", lambda e: self.on_select_added_key_column())
        self._ui_key_list = key_list_list
        key_list_delete_button: Button = Button(key_list_container, text="Remove key", command=self.on_click_remove_key,
                                                background="#ffafaf", activebackground="#bc6262", state="disabled")
        key_list_delete_button.grid_configure(column=0, row=3, sticky="W")
        self._ui_key_delete_button = key_list_delete_button

        output_file_container: Frame = Frame(diff_config_row)
        output_file_container.grid_configure(column=1, row=0, sticky="NSEW", padx=diff_config_row_spacing, pady=diff_config_row_spacing)
        output_file_container.grid_columnconfigure(index=0, weight=1)
        output_file_button: Button = Button(output_file_container, text="Choose diff location", command=self.on_click_choose_destination)
        output_file_button.grid_configure(column=0, row=0, sticky="EW", padx=diff_config_row_spacing)
        output_file_path_label: Label = Label(output_file_container, text="(No destination chosen)")
        output_file_path_label.grid_configure(column=0, row=1, padx=diff_config_row_spacing)
        self._ui_destination_file_label = output_file_path_label



        self._show_button_row()

        self.update_first_file_table_menu()
        self.update_second_file_table_menu()

        window.mainloop()

    # region on_event methods

    def on_click_choose_first_file(self):
        path: str = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")], title="First file")

        if(path == ""):
            return

        self.first_file_path = path
        wb: Workbook = openpyxl.load_workbook(path, data_only=True)
        self.tables_in_first_file = self._get_tables_in_file(path, wb)
        self.column_names_in_tables_from_first_file = self._get_column_names_in_file(wb)
        self._ui_first_file_label.config(text=os.path.basename(path))
        self.update_first_file_table_menu()
        self.update_key_menu()
        self.update_button_row()

    def on_click_choose_second_file(self):
        path: str = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")], title="Second file")

        if(path == ""):
            return

        self.second_file_path = path
        wb: Workbook = openpyxl.load_workbook(path, data_only=True)
        self.tables_in_second_file = self._get_tables_in_file(path, wb)
        self.column_names_in_tables_from_second_file = self._get_column_names_in_file(wb)
        self._ui_second_file_label.config(text=os.path.basename(path))
        self.update_second_file_table_menu()
        self.update_key_menu()
        self.update_button_row()

    def on_choose_first_table(self):
        self.update_selected_first_file_table()
        self.update_button_row()
        self.update_key_menu()

    def on_choose_second_table(self):
        self.update_selected_second_file_table()
        self.update_button_row()
        self.update_key_menu()

    def on_choose_key_column(self):
        key_chooser: Combobox = self._ui_key_list_menu
        key_name: str = key_chooser.get()

        key_chooser.set("Choose a column...")
        self.key_column_names.append(key_name)
        self._ui_key_list.insert("end", key_name)

        self.update_key_menu()
        self.update_key_delete_button()
        self.update_button_row()

    def on_select_added_key_column(self):
        self.update_key_delete_button()

    def on_click_remove_key(self):
        key_list: Listbox = self._ui_key_list
        curselection: tuple = key_list.curselection()

        if(len(curselection) == 0):
            return

        selected_index = curselection[0]
        del self.key_column_names[selected_index]
        key_list.delete(selected_index)
        self.update_key_delete_button()
        self.update_button_row()
        self.update_key_menu()

    def on_click_choose_destination(self):
        path: str = filedialog.asksaveasfilename(confirmoverwrite=True, defaultextension=".xlsx", filetypes=[("Excel file", "*.xlsx")])

        if(path == ""):
            return

        self.destination_file_path = path
        self._ui_destination_file_label.config(text=os.path.basename(path))
        self.update_button_row()

    def on_click_enqueue(self):

        first_table:  TableReference = self.table_selected_from_first_file
        second_table: TableReference = self.table_selected_from_second_file
        diff:         TableDiff      = TableDiff(first_table, second_table, self.destination_file_path, self.key_column_names)

        self.diff_queue.append(diff)
        self.clear_inputs()
        self._switch_to_diff_queue()

    def on_click_create_single_diff(self):
        self._ui_create_diff_button["state"] = "disabled"
        self._ui_enqueue_button["state"] = "disabled"
        first_table: TableReference = self.table_selected_from_first_file
        second_table: TableReference = self.table_selected_from_second_file

        diff: TableDiff = TableDiff(first_table, second_table, self.destination_file_path, self.key_column_names)
        diff.process_and_save()
        self.clear_inputs()

    def on_click_add_to_queue_button(self):
        first_table: TableReference = self.table_selected_from_first_file
        second_table: TableReference = self.table_selected_from_second_file
        diff: TableDiff = TableDiff(first_table, second_table, self.destination_file_path, self.key_column_names)

        self.diff_queue.append(diff)
        self.add_diff_to_queue_listbox(diff)
        self.clear_inputs()

    def on_selected_queued_diff(self):
        self.update_remove_diff_button()

    def on_click_remove_from_queue(self):
        key_list: Listbox = self._ui_diff_queue_listbox
        curselection: tuple = key_list.curselection()

        if(len(curselection) == 0):
            return

        selected_index = curselection[0]
        del self.diff_queue[selected_index]
        key_list.delete(selected_index)
        self.update_remove_diff_button()

        if(len(self.diff_queue) == 0):
            self._switch_to_button_row()

    def on_click_create_diffs(self):
        diffs_to_process = [x for x in self.diff_queue]
        self.clear_queue()

        for diff in diffs_to_process:
            diff.process_and_save()


    #endregion

    #region update UI element methods

    def update_first_file_table_menu(self):
        if(self.first_file_path is None):
            self.table_selected_from_first_file = None
            self._ui_first_file_table_menu.set("Choose a table...")
            self._ui_first_file_table_menu["state"] = "disabled"
            return

        self._ui_first_file_table_menu.config(values=[x.table_name for x in self.tables_in_first_file])
        self._ui_first_file_table_menu["state"] = "readonly"

    def update_second_file_table_menu(self):
        if(self.second_file_path is None):
            self.table_selected_from_second_file = None
            self._ui_second_file_table_menu.set("Choose a table...")
            self._ui_second_file_table_menu["state"] = "disabled"
            return

        self._ui_second_file_table_menu.config(values=[x.table_name for x in self.tables_in_second_file])
        self._ui_second_file_table_menu["state"] = "readonly"

    def update_selected_first_file_table(self):
        if(self.first_file_path is None):
            self._ui_first_file_table_menu.set("Choose a table...")
            self._ui_first_file_table_menu["state"] = "disabled"
            return

        table_name: str = self._ui_first_file_table_menu.get()

        matching_tables: list[TableReference] = [x for x in self.tables_in_first_file if x.table_name == table_name]
        self.table_selected_from_first_file = matching_tables[0] if len(matching_tables) > 0 else None

    def update_selected_second_file_table(self):
        if(self.first_file_path is None):
            self._ui_second_file_table_menu.set("Choose a table...")
            self._ui_second_file_table_menu["state"] = "disabled"
            return

        table_name: str = self._ui_second_file_table_menu.get()

        matching_tables: list[TableReference] = [x for x in self.tables_in_second_file if x.table_name == table_name]
        self.table_selected_from_second_file = matching_tables[0] if len(matching_tables) > 0 else None

    def update_key_menu(self):
        if(self.table_selected_from_first_file is None or self.table_selected_from_second_file is None):
            self._ui_key_list_menu.set("Choose a column...")
            self._ui_key_list_menu["state"] = "disabled"
            return

        column_names_from_first_table \
            = self.column_names_in_tables_from_first_file[self.table_selected_from_first_file.table_name]

        column_names_from_second_table \
            = self.column_names_in_tables_from_second_file[self.table_selected_from_second_file.table_name]

        shared_column_names: list[str] \
            = [x for x in column_names_from_first_table if x in column_names_from_second_table]

        if(len(shared_column_names) == 0):
            self._ui_key_list_menu.set("(No shared columns)")
            self._ui_key_list_menu["state"] = "disabled"
            return

        for column_name in self.key_column_names:
            shared_column_names.remove(column_name)

        if(len(shared_column_names) == 0):
            self._ui_key_list_menu.set("(No remaining shared columns)")
            self._ui_key_list_menu["state"] = "disabled"
            return

        self._ui_key_list_menu.config(values=shared_column_names)
        self._ui_key_list_menu.set("Choose a column...")
        self._ui_key_list_menu["state"] = "readonly"

    def update_key_delete_button(self):
        if(len(self._ui_key_list.curselection()) == 0):
            self._ui_key_delete_button["state"] = "disabled"
            return

        self._ui_key_delete_button["state"] = "normal"

    def update_button_row(self):
        self.update_enqueue_button()
        self.update_create_diff_button()
        self.update_add_to_queue_button()

    def update_enqueue_button(self):
        if(self._ui_enqueue_button is None):
            return

        if(None in [self.first_file_path, self.second_file_path, self.destination_file_path,
                    self.table_selected_from_first_file, self.table_selected_from_second_file]
                or len(self.key_column_names) == 0):

            self._ui_enqueue_button["state"] = "disabled"
            return

        self._ui_enqueue_button["state"] = "normal"

    def update_create_diff_button(self):
        if(self._ui_create_diff_button is None):
            return

        if(None in [self.first_file_path, self.second_file_path, self.destination_file_path,
                    self.table_selected_from_first_file, self.table_selected_from_second_file]
                or len(self.key_column_names) == 0):

            self._ui_create_diff_button["state"] = "disabled"
            return

        self._ui_create_diff_button["state"] = "normal"

    def update_add_to_queue_button(self):
        if(self._ui_add_to_queue_button is None):
            return

        if(None in [self.first_file_path, self.second_file_path, self.destination_file_path,
                    self.table_selected_from_first_file, self.table_selected_from_second_file]
                or len(self.key_column_names) == 0):

            self._ui_add_to_queue_button["state"] = "disabled"
            return

        self._ui_add_to_queue_button["state"] = "normal"

    def update_diff_queue_listbox(self):
        listbox: Listbox = self._ui_diff_queue_listbox
        listbox.delete(0, "end")

        for diff in self.diff_queue:
            self.add_diff_to_queue_listbox(diff)

    def update_remove_diff_button(self):
        if(len(self._ui_diff_queue_listbox.curselection()) == 0):
            self._ui_remove_from_queue_button["state"] = "disabled"
            return

        self._ui_remove_from_queue_button["state"] = "normal"

    # endregion

    def add_diff_to_queue_listbox(self, diff: TableDiff):
        listbox: Listbox = self._ui_diff_queue_listbox
        first_file_name:  str = os.path.basename(diff.first_table_ref.filepath)
        second_file_name: str = os.path.basename(diff.second_table_ref.filepath)
        dest_file_name:   str = os.path.basename(diff.result_filepath)
        listbox.insert("end", f"{first_file_name} & {second_file_name} -> {dest_file_name}")

    def clear_inputs(self):
        self._ui_first_file_label.config(text="(No file selected)")
        self.first_file_path = None
        self.tables_in_first_file = None
        self.column_names_in_tables_from_first_file = None
        self.table_selected_from_first_file = None

        self._ui_second_file_label.config(text="(No file selected)")
        self.second_file_path = None
        self.tables_in_second_file = None
        self.column_names_in_tables_from_second_file = None
        self.table_selected_from_second_file = None

        self.key_column_names = []

        self._ui_destination_file_label.config(text="(No destination chosen)")
        self.destination_file_path = None

        self._ui_key_list.delete(0, "end")

        self.update_first_file_table_menu()
        self.update_second_file_table_menu()
        self.update_key_menu()
        self.update_key_delete_button()
        self.update_button_row()
        self.update_selected_first_file_table()
        self.update_selected_second_file_table()

    def clear_queue(self):
        self.diff_queue.clear()

        if(self._ui_diff_queue is not None):
            self._switch_to_button_row()

    def _hide_button_row(self):
        if(self._ui_button_row_without_queue is None):
            return

        self._ui_button_row_without_queue.pack_forget()
        self._ui_button_row_without_queue = None
        self._ui_enqueue_button           = None
        self._ui_create_diff_button       = None

    def _hide_diff_queue(self):
        if(self._ui_diff_queue is None):
            return

        self._ui_diff_queue.pack_forget()
        self._ui_diff_queue = None

    def _show_button_row(self):
        action_button_row: Frame = Frame(self._ui_window)
        action_button_row.pack_configure(side="bottom", fill="x", pady=(20, 0))
        enqueue_button: Button = Button(action_button_row, text="Enqueue", command=self.on_click_enqueue)
        enqueue_button.pack_configure(side="left", padx=5, pady=5)
        enqueue_button["state"] = "disabled"
        create_diff_button: Button = Button(action_button_row, text="Create diff", command=self.on_click_create_single_diff)
        create_diff_button.pack_configure(side="right", fill="x", expand=True, padx=5, pady=5)
        self._ui_enqueue_button           = enqueue_button
        self._ui_create_diff_button       = create_diff_button
        self._ui_button_row_without_queue = action_button_row

        self.update_button_row()

    def _show_diff_queue(self):
        diff_queue: Frame = Frame(self._ui_window)
        diff_queue.pack_configure(side="bottom", fill="both", expand=True)

        add_to_queue_button_container: Frame = Frame(diff_queue)
        add_to_queue_button_container.pack_configure(side="top", fill="x")
        add_to_queue_button: Button = Button(add_to_queue_button_container, text="+ Add to queue",
                                             command=self.on_click_add_to_queue_button)
        add_to_queue_button.pack_configure(fill="both", padx=5, pady=5)
        self._ui_add_to_queue_button = add_to_queue_button

        queue_list: Listbox = Listbox(diff_queue, height=2)
        queue_list.pack_configure(fill="both", expand=True, padx=5)
        queue_list.bind("<<ListboxSelect>>", lambda e: self.on_selected_queued_diff())
        self._ui_diff_queue_listbox = queue_list

        diff_queue_bottom_button_row: Frame = Frame(diff_queue)
        diff_queue_bottom_button_row.pack_configure(side="bottom", fill="x")
        remove_diff_from_queue_button: Button = Button(diff_queue_bottom_button_row, text="Remove",
                                                       background="#ffafaf", activebackground="#bc6262",
                                                       command=self.on_click_remove_from_queue)
        remove_diff_from_queue_button.pack_configure(fill="none", padx=5, pady=5, side="left")
        remove_diff_from_queue_button["state"] = "disabled"
        self._ui_remove_from_queue_button = remove_diff_from_queue_button
        create_diffs_from_queue_button: Button = Button(diff_queue_bottom_button_row, text="Create diffs",
                                                        command=self.on_click_create_diffs)
        create_diffs_from_queue_button.pack_configure(fill="x", expand=True, padx=5, pady=5, side="right")
        self._ui_create_diffs_from_queue_button = create_diffs_from_queue_button

        self._ui_diff_queue          = diff_queue

        self.update_diff_queue_listbox()

    def _switch_to_button_row(self):
        self._hide_button_row()
        self._hide_diff_queue()
        self._show_button_row()

    def _switch_to_diff_queue(self):
        self._hide_button_row()
        self._hide_diff_queue()
        self._show_diff_queue()

    def _get_tables_in_file(self, filepath: str, wb: Workbook) -> list[TableReference]:
        result: list[TableReference] = []
        sheet_names = wb.sheetnames

        for sheet_name in sheet_names:
            sheet: Worksheet = wb.get_sheet_by_name(sheet_name)

            table_name: str

            for table_name in sheet.tables.keys():
                result.append(TableReference(filepath, sheet_name, table_name))

        return result

    def _get_column_names_in_file(self, wb: Workbook) -> dict[str, list[str]]:
        result: dict[str, list[str]] = {}

        for sheet in wb.worksheets:
            name: str

            for name in sheet.tables.keys():
                table: Table = sheet.tables[name]
                result[name] = table.column_names

        return result
