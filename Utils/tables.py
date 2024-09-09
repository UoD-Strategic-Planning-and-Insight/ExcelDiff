"""
Contains a class with improved support for accessing and modifying tables in an Excel file via OpenPyxl.
"""

from typing import Generator

from openpyxl.cell import Cell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet

import Utils


class TableUtil:
    """
    Utility wrapper class for openpyxl's Table class, adding some functionality or making some functionality more easily
    accessible.
    """

    filepath: str
    """The filepath of the workbook (.xlsx file) the table this represents is in."""

    source_workbook: Workbook
    """The workbook (.xlsx file) the table this represents is in."""

    source_worksheet: Worksheet
    """The worksheet the table this represents is in."""

    source_table: Table
    """The extended table."""

    min_col_name: str = ""
    """
    The name of the earliest column in the table. This is the column number as it appears in Excel- e.g. "AA" would be
    the 27th column.
    """

    max_col_name: str = ""
    """
    The name of the last column in the table. This is the column number as it appears in Excel- e.g. "AA" would be
    the 27th column.
    """

    min_col: int = -1
    """The column number of the earliest column in the table."""

    max_col: int = -1
    """The column number of the last column in the table."""

    min_row: int = -1
    """The row number of the earliest row in the table."""

    max_row: int = -1
    """The row number of the last row in the table."""

    ref: str
    """The table range reference string. e.g. "A2:C20"."""

    def __init__(self, filepath: str, workbook: Workbook, worksheet: Worksheet, table: Table):
        """
        Wraps an existing openpyxl table object in a new table util class.
        :param worksheet: The worksheet the table this wraps is in.
        :param table: The table this is to wrap.
        """

        self.filepath         = filepath
        self.source_workbook  = workbook
        self.source_worksheet = worksheet
        self.source_table     = table
        self.ref              = table.ref
        splits: list[str]     = table.ref.split(":")

        top_left:     str = splits[0]
        bottom_right: str = splits[1]

        for i in range(len(top_left)):
            c: str = top_left[i]

            if(c.isdigit()):
                self.min_col_name = top_left[:i]
                self.min_col = Utils.convert_alphabetic_number_to_int(self.min_col_name)
                self.min_row = int(top_left[i:])
                break

        for i in range(len(bottom_right)):
            c: str = bottom_right[i]

            if(c.isdigit()):
                self.max_col_name = bottom_right[:i]
                self.max_col = Utils.convert_alphabetic_number_to_int(self.max_col_name)
                self.max_row = int(bottom_right[i:])
                break

    def iter_rows(self) -> Generator[tuple[Cell, ...], None, None]:
        """
        Provides an iterator over the rows of the table, excluding any header rows.
        :return: An iterator over the rows of the table, as tuples of cells as provided by `Worksheet.iter_rows`.
        """

        return self.source_worksheet.iter_rows(self.min_row + self.source_table.headerRowCount,
                                               self.max_row,
                                               self.min_col,
                                               self.max_col)

    def iter_rows_with_column_names(self) -> Generator[dict[str, Cell], None, None]:
        """
        Provides an iterator over the rows of the table, excluding any header rows.
        :return: An iterator over the rows of the table, as dictionaries of cells where the key is the column header.
                 Note that this assumes the table has at least one header row, and assumes that the final header row is
                 the one containing the column titles.
        """

        col_names: list[str] = []

        col_name_iterator = self.source_worksheet.iter_rows(self.min_row + self.source_table.headerRowCount - 1,
                                                            self.min_row + self.source_table.headerRowCount - 1,
                                                            self.min_col,
                                                            self.max_col)

        row_iterator = self.source_worksheet.iter_rows(self.min_row + self.source_table.headerRowCount,
                                                       self.max_row,
                                                       self.min_col,
                                                       self.max_col)

        for row in col_name_iterator:
            col_names = []

            for cell in row:
                col_names.append(str(cell.value))

        for row in row_iterator:
            cells: dict[str, Cell] = {}

            for i in range(len(col_names)):
                cells[col_names[i]] = row[i]

            yield cells

    def add_row(self, values: list[any]):
        """
        Adds a row to the table this represents. This extends the table down one row in the sheet it's in.
        :param values: The values to populate the table rows with.
        """

        self.max_row += 1
        self.ref = f"{self.min_col_name}{self.min_row}:{self.max_col_name}{self.max_row}"
        self.source_table.ref = self.ref

        for i in range(len(values)):
            self.source_worksheet.cell(self.max_row, self.min_col + i).value = values[i]

    def add_column(self, column_name: str):
        """
        Adds a column to the table this represents. This extends the table one column to the right in the sheet it's in.
        :param column_name: The name of the column. This will appear in the column header.
        """

        self.source_worksheet.cell(self.min_row, self.max_col + 1).value = column_name

        self.max_col += 1
        self.max_col_name = Utils.convert_int_to_alphabetic_number(self.max_col)
        self.ref = f"{self.min_col_name}{self.min_row}:{self.max_col_name}{self.max_row}"
        self.source_table.ref = self.ref

    def get_column_names(self) -> list[str]:
        """
        Gets the names of the columns in this table.
        :return: The names of the columns in this table. That is, the text appearing in the header of each column, in
        order of column number.
        """

        col_names: list[str] = []

        col_name_iterator = self.source_worksheet.iter_rows(self.min_row + self.source_table.headerRowCount - 1,
                                                            self.min_row + self.source_table.headerRowCount - 1,
                                                            self.min_col,
                                                            self.max_col)

        for row in col_name_iterator:
            col_names = []

            for cell in row:
                col_names.append(str(cell.value))

        return col_names

    def get_bottom_row(self) -> dict[str, Cell]:
        """
        Gets the last row in the table.
        :return: The contents of the last row in the table as a dictionary, where the keys are the column names (i.e.
                 the text that appears in the column's header cell) and the values are the contents of the cells in the
                 corresponding columns of the row.
        """

        col_names = self.get_column_names()
        result: dict[str, Cell] = {}

        for i in range(self.max_col - self.min_col + 1):
            result[col_names[i]] = self.source_worksheet.cell(self.max_row, self.min_col + i)

        return result

    def get_row(self, row_number: int) -> dict[str, Cell]:
        """
        Gets the nth row in the table. Note that the row number here is zero-indexed, and begins at the first data row.
        (under the header row)
        :param row_number: The zero-indexed number of the row to get.
        :return: The contents of the row at the given row number in the table as a dictionary, where the keys are the
                 column names (i.e. the text that appears in the column's header cell) and the values are the contents
                 of the cells in the corresponding columns of the row.
        """

        # When writing documentation, note that the row number is zero-indexed row number *within the table*. The row
        # number 0 here may have the row number 2 in the sheet. (since the sheet is 1-indexed + the headers take up a
        # row)

        col_names = self.get_column_names()
        result: dict[str, Cell] = {}

        for i in range(self.max_col - self.min_col + 1):
            row_no_in_sheet: int = self.min_row + row_number + self.source_table.headerRowCount
            col_no_in_sheet: int = self.min_col + i
            result[col_names[i]] = self.source_worksheet.cell(row_no_in_sheet, col_no_in_sheet)

        return result

    def has_column(self, column_name: str) -> bool:
        """
        Gets whether a column exists in this table with a given name.
        :param column_name: The name of the column to look for.
        :return: True if any column has the given name - that is, any column's header cell matches the given text.
                 Otherwise, false.
        """

        return column_name in self.get_column_names()

    def save(self):
        """
        Saves the table, committing any changes made to the file.

        Note that this saves the entire .xlsx file the table is in - if any changes have been made elsewhere, this will
        save them as well.
        """

        self.source_workbook.save(self.filepath)
