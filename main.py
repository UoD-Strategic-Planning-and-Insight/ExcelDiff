from dataclasses import dataclass

import openpyxl
from openpyxl.cell import Cell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet

import Utils
import Utils.boilerplate as bp
from Utils.tables import TableUtil


@dataclass
class CellDifference:
    column_name: str
    value1: str
    value2: str


@dataclass
class RowDifference:
    keys: dict[str, str]
    cell_differences: list[CellDifference]


@dataclass
class DiffRequest:
    old_path: str
    new_path: str
    diff_path: str


@dataclass
class ColumnContent:
    column_name: str
    column_cell_values: list[any]


row_numbers_for_key_sets_in_old: dict[str, int] = {}
row_numbers_for_key_sets_in_new: dict[str, int] = {}


def main():

    # The following is an example

    key_columns: list[str] = ["key1", "key2", "key3"]

    sheet_name: str = "Sheet1"
    table_name: str = "Table1"

    folder_path: str = (r"C:\Users\yourusername\Desktop\Example folder\\")

    diff_requests: list[DiffRequest] = \
        [
            DiffRequest(folder_path + "", folder_path + "", folder_path + "")
        ]

    produce_diffs(key_columns, sheet_name, table_name, diff_requests)


def produce_diffs(key_columns: list[str], sheet_name: str, table_name: str, diff_requests: list[DiffRequest]):
    for diff_request in diff_requests:
        print(f"\nAbout to diff:\n    {diff_request.old_path}\n  & {diff_request.new_path}\n -> {diff_request.diff_path}")

        produce_diff(key_columns, sheet_name, table_name,
                     diff_request.old_path, diff_request.new_path, diff_request.diff_path)


def produce_diff(key_columns:    list[str],
                 sheet_name:     str,
                 table_name:     str,
                 old_file_path:  str,
                 new_file_path:  str,
                 diff_file_path: str):

    old_tbl = bp.load_table(old_file_path, sheet_name, table_name)
    new_tbl = bp.load_table(new_file_path, sheet_name, table_name)

    different_rows: list[RowDifference] = []
    rows_only_in_1: list[dict[str, Cell]] = []
    rows_only_in_2: list[dict[str, Cell]] = []

    print("Prepopulating the caches")

    pre_populate_cache(old_tbl, key_columns, row_numbers_for_key_sets_in_old)
    pre_populate_cache(new_tbl, key_columns, row_numbers_for_key_sets_in_new)

    print("About to check diffs.")

    counter: int = 0

    for row in old_tbl.iter_rows_with_column_names():
        counter += 1

        if(counter % 50 == 0):
            print(f"  - Row {counter}")

        keys: dict[str, str] = {}

        for key_column_name in key_columns:
            keys[key_column_name] = str(row[key_column_name].value)

        matching_new_row: dict[str, Cell] = get_row_with_keys(new_tbl, keys, row_numbers_for_key_sets_in_new)

        if(matching_new_row is None):
            rows_only_in_1.append(row)
            continue

        row_diff: list[CellDifference] = get_differences_between_rows(row, matching_new_row)

        if(len(row_diff) != 0):
            different_rows.append(RowDifference(keys, row_diff))

    print("About to check for rows only in 2nd table.")

    counter = 0

    for row in new_tbl.iter_rows_with_column_names():
        counter += 1

        if (counter % 50 == 0):
            print(f"  - Row {counter}")

        keys: dict[str, str] = {}

        for key_column_name in key_columns:
            keys[key_column_name] = str(row[key_column_name].value)

        matching_old_row: dict[str, Cell] = get_row_with_keys(old_tbl, keys, row_numbers_for_key_sets_in_old)

        if(matching_old_row is None):
            rows_only_in_2.append(row)

    for row in different_rows:
        print(f"{row}")

    print("About to check for columns only in 1st table.")

    col_names_1 = old_tbl.get_column_names()
    col_names_2 = new_tbl.get_column_names()

    cols_only_in_1: list[ColumnContent] = []

    for col_name in col_names_1:
        if col_name not in col_names_2:
            col_vals: list[any] = []

            for row in old_tbl.iter_rows_with_column_names():
                col_vals.append(row[col_name].value)

            cols_only_in_1.append(ColumnContent(col_name, col_vals))

    key_cols_in_1: list[ColumnContent] | None = None

    if(len(cols_only_in_1) != 0):
        key_cols_in_1 = []

        for col_name in key_columns:
            col_vals: list[any] = []

            for row in old_tbl.iter_rows_with_column_names():
                col_vals.append(row[col_name].value)

            key_cols_in_1.append(ColumnContent(col_name, col_vals))

    print("About to check for columns only in 2nd table.")

    cols_only_in_2: list[ColumnContent] = []

    for col_name in col_names_2:
        if(col_name not in col_names_1):
            col_vals: list[any] = []

            for row in new_tbl.iter_rows_with_column_names():
                col_vals.append(row[col_name].value)

            cols_only_in_2.append(ColumnContent(col_name, col_vals))

    key_cols_in_2: list[ColumnContent] | None = None

    if(len(cols_only_in_2) != 0):
        key_cols_in_2 = []

        for col_name in key_columns:
            col_vals: list[any] = []

            for row in new_tbl.iter_rows_with_column_names():
                col_vals.append(row[col_name].value)

            key_cols_in_2.append(ColumnContent(col_name, col_vals))

    save_diffs_to_file(different_rows,
                       rows_only_in_1,
                       rows_only_in_2,
                       key_cols_in_1,
                       key_cols_in_2,
                       cols_only_in_1,
                       cols_only_in_2,
                       diff_file_path)


def replace_quote_in_str(source: str) -> str:
    quote_replacement_base: str = "quote$"
    quote_replacement: str = "[quote$1]"
    quote_replacement_int: int = 1

    while(quote_replacement in source):
        quote_replacement_int += 1
        quote_replacement = f"[{quote_replacement_base}{quote_replacement_int}]"

    return source.replace("\"", quote_replacement)


def dict_to_str(dictionary: dict[str, str]) -> str:
    # TODO: Note in documentation that this replaces quotes in the key and value to guarantee uniqueness is retained.

    sorted_keys: list[str] = sorted(dictionary.keys())
    keys_and_vals_as_strs: list[str] = []

    for key in sorted_keys:
        adjusted_key   = replace_quote_in_str(key)
        adjusted_value = replace_quote_in_str(dictionary[key])
        keys_and_vals_as_strs.append(f"\"{adjusted_key}\": \"{adjusted_value}\"")

    return "{" + (", ".join(keys_and_vals_as_strs)) + "}"


def pre_populate_cache(tbl: TableUtil, key_cols: list[str], cache: dict[str, int]):
    row_no: int = -1

    for row in tbl.iter_rows_with_column_names():
        keys: dict[str, str] = {}
        row_no += 1

        for key_col in key_cols:
            keys[key_col] = str(row[key_col].value)

        key_str: str = dict_to_str(keys)
        cache[key_str] = row_no


def get_row_with_keys(tbl: TableUtil, key_dict: dict[str, str], cache: dict[str, int]) \
        -> dict[str, Cell] | None:

    key_str: str = dict_to_str(key_dict)
    cached_row_no: int | None = cache.get(key_str)

    if(cached_row_no is not None):
        return tbl.get_row(cached_row_no)

    row: dict[str, Cell]
    row_no: int = -1

    for row in tbl.iter_rows_with_column_names():
        matches: bool = True
        row_no += 1

        for k, v in key_dict.items():
            if(str(row[k].value) != str(v)):
                matches = False
                break

        if(matches):
            cache[key_str] = row_no
            return row

    return None


def get_differences_between_rows(row1: dict[str, Cell], row2: dict[str, Cell]) -> list[CellDifference]:
    result: list[CellDifference] = []

    for k, v1 in row1.items():
        v2: Cell | None = row2.get(k)

        if(v2 is None):
            continue

        v1val = str(v1.value).strip() if v1.value is not None else ""
        v2val = str(v2.value).strip() if v2.value is not None else ""

        if(v1val != v2val):
            result.append(CellDifference(k, v1val, v2val))

    return result


def save_diffs_to_file(diffs:          list[RowDifference],
                       rows_only_in_1: list[dict[str, Cell]],
                       rows_only_in_2: list[dict[str, Cell]],
                       key_cols_in_1:  list[ColumnContent] | None,
                       key_cols_in_2:  list[ColumnContent] | None,
                       cols_only_in_1: list[ColumnContent],
                       cols_only_in_2: list[ColumnContent],
                       filepath:       str):
    wb = openpyxl.Workbook()

    add_diffs_sheet(wb, filepath, diffs)
    add_rows_only_in_one_sheet(wb, filepath, rows_only_in_1, 1)
    add_rows_only_in_one_sheet(wb, filepath, rows_only_in_2, 2)

    add_columns_only_in_one_sheet(wb, filepath, key_cols_in_1, cols_only_in_1, 1)
    add_columns_only_in_one_sheet(wb, filepath, key_cols_in_2, cols_only_in_2, 2)

    wb.save(filepath)


def add_diffs_sheet(workbook: Workbook, filepath: str, diffs: list[RowDifference]):
    if(len(diffs) == 0):
        return

    workbook.create_sheet("Diff", 0)
    sheet = workbook.get_sheet_by_name("Diff")

    keys = [x for x in diffs[0].keys.keys()]

    for i in range(len(keys)):
        sheet.cell(1, i + 1).value = keys[i]

    table = Table(displayName="DiffTable", ref=f"A1:{Utils.convert_int_to_alphabetic_number(len(keys))}1")
    sheet.add_table(table)

    tbl_diff = TableUtil(filepath, workbook, sheet, table)

    for diff in diffs:
        tbl_diff.add_row([])
        row = tbl_diff.get_bottom_row()
        for k, v in diff.keys.items():
            row[k].value = v

        for cell_diff in diff.cell_differences:
            col_name_1 = cell_diff.column_name + " * 1"
            col_name_2 = cell_diff.column_name + " * 2"
            bottom_row_changed: bool = False

            if(not tbl_diff.has_column(col_name_1)):
                tbl_diff.add_column(col_name_1)
                tbl_diff.add_column(col_name_2)
                bottom_row_changed = True

            if(bottom_row_changed):
                row = tbl_diff.get_bottom_row()

            row[col_name_1].value = cell_diff.value1
            row[col_name_2].value = cell_diff.value2


def add_rows_only_in_one_sheet(workbook: Workbook,
                               filepath: str,
                               rows_only_in_one: list[dict[str, Cell]],
                               index: int):

    if(len(rows_only_in_one) == 0):
        return

    sheet_name: str = f"Rows only in {index}"
    workbook.create_sheet(sheet_name, index)
    sheet = workbook.get_sheet_by_name(sheet_name)

    column_names: list[str] = [x for x in rows_only_in_one[0].keys()]

    for i in range(len(column_names)):
        sheet.cell(1, i + 1).value = column_names[i]

    table = Table(displayName=f"OnlyIn{index}", ref=f"A1:{Utils.convert_int_to_alphabetic_number(len(column_names))}1")
    sheet.add_table(table)

    tbl_util = TableUtil(filepath, workbook, sheet, table)

    for row in rows_only_in_one:
        tbl_util.add_row([])
        bottom_row = tbl_util.get_bottom_row()

        for k, v in row.items():
            bottom_row[k].value = v.value

def add_columns_only_in_one_sheet(workbook: Workbook,
                                  filepath: str,
                                  key_columns: list[ColumnContent],
                                  columns_only_in_one: list[ColumnContent],
                                  index: int):

    if(len(columns_only_in_one) == 0):
        return

    sheet_name: str = f"Columns only in {index}"
    workbook.create_sheet(sheet_name, index)
    sheet = workbook.get_sheet_by_name(sheet_name)

    row_count: int = len(columns_only_in_one[0].column_cell_values)

    for i in range(len(key_columns)):
        col_content: ColumnContent = key_columns[i]
        sheet.cell(1, i + 1).value = col_content.column_name

        for j in range(len(col_content.column_cell_values)):
            sheet.cell(j + 2, i + 1).value = col_content.column_cell_values[j]

    table = Table(displayName=f"ColOnlyIn{index}",
                  ref=f"A1:{Utils.convert_int_to_alphabetic_number(len(key_columns))}{row_count + 1}")

    sheet.add_table(table)

    tbl_util = TableUtil(filepath, workbook, sheet, table)

    for col in columns_only_in_one:
        tbl_util.add_column(col.column_name, col.column_cell_values)


if __name__ == '__main__':
    main()
